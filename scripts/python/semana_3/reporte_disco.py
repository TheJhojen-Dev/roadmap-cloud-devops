import os
import subprocess
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment


def get_disk_usage(target_dir: str) -> pd.DataFrame:
    cmd = ["du", "-d", "1", "-k", target_dir]
    result = subprocess.run(cmd, capture_output=True, text=True, check=True)

    target_abs = os.path.abspath(target_dir)
    records = []

    for line in result.stdout.strip().split("\n"):
        parts = line.split("\t")
        if len(parts) != 2:
            continue

        kb, path = parts
        path_abs = os.path.abspath(path)

        if path_abs == target_abs:
            continue

        folder_name = os.path.basename(path)
        if not folder_name:
            continue

        mb = round(int(kb) / 1024, 2)
        if mb > 0:
            records.append({"Carpeta": folder_name, "Espacio (MB)": mb})

    df = pd.DataFrame(records)
    return df.sort_values(by="Espacio (MB)", ascending=False).reset_index(
        drop=True
    )


def format_excel_report(file_path: str, df_len: int) -> None:
    wb = load_workbook(file_path)
    ws = wb["Uso_Disco"]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")

    chart_limit = min(df_len + 1, 16)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Top 15 - Consumo de Espacio (MB)"
    chart.y_axis.title = "Megabytes (MB)"
    chart.x_axis.title = "Carpetas"
    chart.width = 20
    chart.height = 12
    chart.legend = None

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=chart_limit)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=chart_limit)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "D2")
    wb.save(file_path)


def generate_report(output_file: str = "Reporte_Espacio_Disco.xlsx") -> None:
    target_dir = os.path.expanduser("~")
    df = get_disk_usage(target_dir)

    df.to_excel(output_file, index=False, sheet_name="Uso_Disco")
    format_excel_report(output_file, len(df))
    print(f"Reporte generado exitosamente: {output_file}")


if __name__ == "__main__":
    generate_report()
